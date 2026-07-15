#!/usr/bin/env python3
"""Refresh the companion literature-audit workbook with execution evidence."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.properties import CalcProperties


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = (
    ROOT
    / "outputs"
    / "literature_audit_2026-07-15"
    / "AI_NeuroOnco_Literature_Audit_and_Index.xlsx"
)


def read_json(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8")) if path.exists() else {}


def main() -> None:
    wb = load_workbook(WORKBOOK)
    if "Research Status" in wb.sheetnames:
        del wb["Research Status"]
    ws = wb.create_sheet("Research Status", 1)

    navy = "17365D"
    blue = "2F75B5"
    pale_blue = "D9EAF7"
    pale_green = "E2F0D9"
    pale_red = "FCE4D6"
    grey = "E7E6E6"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = "Research Execution Status and Evidence Index"
    ws["A1"].font = Font(size=18, bold=True, color=white)
    ws["A1"].fill = PatternFill("solid", fgColor=navy)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells("A3:F3")
    ws["A3"] = (
        "This sheet links the literature audit to generated study evidence. "
        "It does not convert missing external or expert evidence into claims."
    )
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws["A3"].fill = PatternFill("solid", fgColor=pale_blue)

    headers = ["Domain", "Status", "Evidence", "Key finding", "Manuscript use", "Remaining boundary"]
    for col, value in enumerate(headers, start=1):
        cell = ws.cell(5, col, value)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border

    dataset = read_json(ROOT / "data" / "manifests" / "dataset_audit.json")
    readiness = read_json(ROOT / "outputs" / "research_readiness.json")
    brisc = read_json(
        ROOT / "data" / "external_manifests" / "brisc_overlap_audit" / "overlap_audit.json"
    )
    bdneuro = read_json(
        ROOT
        / "data"
        / "external_manifests"
        / "bdneuro_mri_v7_overlap_audit"
        / "overlap_audit.json"
    )
    rag = read_json(ROOT / "outputs" / "rag_evaluation" / "retrieval_metrics.json")
    custom = read_json(ROOT / "outputs" / "metrics" / "custom_cnn" / "research_metrics.json")
    efficient = read_json(ROOT / "outputs" / "metrics" / "efficientnet" / "research_metrics.json")
    efficient_run = read_json(
        ROOT / "outputs" / "models" / "efficientnet" / "run_summary.json"
    )
    comparison = read_json(
        ROOT / "outputs" / "model_comparison" / "model_comparison.json"
    )
    xai = read_json(ROOT / "outputs" / "xai_evaluation" / "xai_metrics.json")

    def metric_text(metrics: dict) -> str:
        if not metrics:
            return "Protocol-conformant run not complete"
        ci = metrics.get("confidence_intervals", {}).get("f1_macro", {})
        return (
            f"Internal macro-F1 {metrics.get('f1_macro', 0):.3f} "
            f"(95% CI {ci.get('lower_95', 0):.3f}–{ci.get('upper_95', 0):.3f}); "
            f"accuracy {metrics.get('accuracy', 0):.3f}"
        )

    rag_agg = rag.get("aggregate", {})
    rows = [
        (
            "Development cohort",
            "Complete" if dataset.get("leakage_free") else "Blocked",
            "data/manifests/dataset_audit.json",
            f"{dataset.get('records', 0):,} retained; {dataset.get('excluded_conflicting_record_count', 0)} conflicting records excluded; no final cross-split hash/group reuse",
            "Methods and cohort results",
            "No patient identifiers; residual same-patient overlap cannot be excluded",
        ),
        (
            "Core literature verification",
            "Partial",
            "research/FULL_TEXT_EXTRACTION.md",
            "9/10 shortlisted records full-text extracted by a first reviewer; two misaligned binary studies removed from the four-class core",
            "Introduction, comparison and limitations",
            "Independent second-reviewer extraction remains required",
        ),
        (
            "Data quality and team handoff",
            "Complete",
            "research/TRACK1_TRACK2_HANDOFF.md; outputs/data_quality/data_quality_report.json",
            "All 23 question-set items answered from evidence; class, dimension, intensity-sample and representative-image outputs generated",
            "Methods, supplement and team handoff",
            "Patient/acquisition fields remain unavailable rather than inferred",
        ),
        ("Custom CNN", "Complete" if custom else "Running / pending", "outputs/metrics/custom_cnn/research_metrics.json", metric_text(custom), "Internal comparator only", "Do not interpret as external generalization"),
        (
            "EfficientNet-B0",
            "Evaluated / resource-constrained" if efficient else "Pending",
            "outputs/metrics/efficientnet/research_metrics.json; outputs/models/efficientnet/run_summary.json",
            metric_text(efficient),
            "Internal comparator with explicit deviation",
            (
                "Stopped after 9 epochs; not protocol-conformant; primary external endpoint also requires an eligible cohort"
                if efficient_run and not efficient_run.get("protocol_conformant")
                else "Primary external endpoint still requires an eligible cohort"
            ),
        ),
        (
            "Paired internal model comparison",
            "Complete" if comparison else "Pending",
            "outputs/model_comparison/model_comparison.json",
            (
                "Grouped paired-bootstrap differences complete; image-level exact McNemar is exploratory"
                if comparison
                else "Requires both per-image locked-test prediction tables"
            ),
            "Internal model-comparison results",
            "Internal source test only; groups are not verified patients",
        ),
        (
            "External candidate: BRISC",
            "Rejected",
            "data/external_manifests/brisc_overlap_audit/overlap_audit.json",
            f"{brisc.get('exact_overlap_candidate_records', 0):,}/6,000 exact-overlap records ({100 * brisc.get('candidate_exact_overlap_fraction', 0):.1f}%)",
            "Candidate-audit result; masks only for internal-test XAI",
            "Not independent external validation",
        ),
        (
            "External candidate: BDNeuro-MRI v7",
            "Rejected",
            "data/external_manifests/bdneuro_mri_v7_overlap_audit/overlap_audit.json",
            f"{bdneuro.get('exact_overlap_candidate_records', 0):,}/{bdneuro.get('candidate_records', 0):,} exact-overlap records ({100 * bdneuro.get('exact_overlap_fraction', 0):.1f}%)",
            "Candidate-audit result",
            "Not independent; distributed consent/license text also incomplete",
        ),
        (
            "Quantitative Grad-CAM",
            "Complete" if xai else "Pending model completion",
            "outputs/xai_evaluation/xai_metrics.json",
            f"{xai.get('cases', 0):,} locked internal-test mask mappings evaluated" if xai else "Repeatability, randomization, IoU and pointing-game runner ready",
            "Ancillary internal-test XAI",
            "Blinded domain-expert review remains required",
        ),
        (
            "Literature RAG",
            "Preliminary",
            "outputs/rag_evaluation/retrieval_metrics.json",
            f"10 cases: recall@5 {rag_agg.get('recall_at_k', 0):.3f}, hit@5 {rag_agg.get('hit_at_k', 0):.3f}, MRR {rag_agg.get('reciprocal_rank', 0):.3f}",
            "Engineering appendix or separate paper",
            "Benchmark is not independently expert-reviewed",
        ),
        (
            "Paper readiness",
            "Ready" if readiness.get("paper_ready") else "Not ready",
            "outputs/research_readiness.json",
            f"Implementation {100 * readiness.get('implementation_readiness_fraction', 0):.0f}%; core empirical {100 * readiness.get('empirical_readiness_fraction', 0):.0f}%",
            "Controls claim promotion",
            "See machine-readable blocking prerequisites",
        ),
    ]
    for row_index, values in enumerate(rows, start=6):
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row_index, col, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        status = values[1]
        ws.cell(row_index, 2).fill = PatternFill(
            "solid",
            fgColor=pale_green if status in {"Complete", "Ready"} else pale_red if status in {"Rejected", "Blocked", "Not ready"} else grey,
        )

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{5 + len(rows)}"
    widths = [28, 20, 52, 58, 36, 48]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width
    for row in range(6, 6 + len(rows)):
        ws.row_dimensions[row].height = 64

    missing = wb["Missing Anchors"]
    for row in range(6, missing.max_row + 1):
        if missing.cell(row, 1).value == "BRISC 2026":
            missing.cell(row, 6).value = "Audited: reject as external; retain masks for internal-test XAI only"
    if not any(missing.cell(row, 1).value == "BDNeuro-MRI 2026" for row in range(6, missing.max_row + 1)):
        row = missing.max_row + 1
        values = [
            "BDNeuro-MRI 2026",
            "Bangladeshi four-class MRI dataset",
            "External candidate audit",
            "Current hospital-labelled candidate; independent audit found material exact reuse and incomplete distributed consent/license text",
            "https://data.mendeley.com/datasets/zwr4ntf94j/7",
            "Audited: reject as independent external validation",
        ]
        for col, value in enumerate(values, start=1):
            cell = missing.cell(row, col, value)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        missing.row_dimensions[row].height = 58

    # Persist first-reviewer full-text verification and metadata corrections in
    # the companion index. The original tracker remains untouched.
    audit_summary = wb["Audit Summary"]
    paper_index = wb["Paper Index"]
    full_text_ids = {10, 11, 12, 13, 14, 16, 31, 32, 34}
    for row in range(6, paper_index.max_row + 1):
        record_id = paper_index.cell(row, 1).value
        if record_id in full_text_ids:
            paper_index.cell(row, 7).value = "Full-text verified (first reviewer)"
            paper_index.cell(row, 18).value = (
                "First-reviewer extraction complete; independent second review pending"
            )
            paper_index.cell(row, 19).value = (
                "Use only with limitations extracted in research/FULL_TEXT_EXTRACTION.md"
            )
        elif record_id == 15:
            paper_index.cell(row, 7).value = "Abstract + publisher record verified"
            paper_index.cell(row, 18).value = (
                "Detailed full-PDF extraction and independent second review pending"
            )
        if record_id == 13:
            paper_index.cell(row, 17).value = "Supporting – binary classifier"
            paper_index.cell(row, 20).value = "No"
            paper_index.cell(row, 19).value = (
                "Remove from four-class core shortlist; retain as binary transfer-learning context"
            )
        elif record_id == 14:
            paper_index.cell(row, 3).value = 2025
        elif record_id == 16:
            paper_index.cell(row, 3).value = 2021
            paper_index.cell(row, 17).value = "Supporting – binary classifier"
            paper_index.cell(row, 20).value = "No"
            paper_index.cell(row, 19).value = (
                "Remove from four-class core shortlist; correct publication year to 2021"
            )
        elif record_id == 32:
            paper_index.cell(row, 13).value = "10.3390/diagnostics12081850"
            paper_index.cell(row, 16).value = "https://doi.org/10.3390/diagnostics12081850"
            paper_index.cell(row, 22).value = "DOI"
        elif record_id == 34:
            paper_index.cell(row, 3).value = 2025

    for row in range(5, audit_summary.max_row + 1):
        if audit_summary.cell(row, 1).value == "Full-text extracted":
            audit_summary.cell(row, 2).value = 9
        elif audit_summary.cell(row, 1).value == "Core shortlist":
            audit_summary.cell(row, 2).value = 8

    priority_actions = [
        ("Obtain a genuinely independent external cohort with compatible labels and auditable provenance.", "P0", "Data", "Before submission"),
        ("Complete blinded neuroradiology/neuro-oncology review of the prespecified Grad-CAM cases.", "P1", "XAI", "Before XAI claims"),
        ("Complete independent second-reviewer extraction and adjudication of the ten-paper queue.", "P1", "Literature", "Before submission"),
        ("Obtain patient/acquisition metadata for patient-level and subgroup claims, or retain the stated non-estimable boundary.", "P1", "Data", "Before broader claims"),
        ("Obtain domain-expert review of the RAG benchmark or keep RAG ancillary/separate.", "P2", "RAG", "If retained"),
    ]
    for offset, (action, priority, domain, timing) in enumerate(priority_actions, start=12):
        audit_summary.cell(offset, 5).value = action
        audit_summary.cell(offset, 6).value = priority
        audit_summary.cell(offset, 7).value = domain
        audit_summary.cell(offset, 8).value = timing

    paper_plan = wb["Paper Plan"]
    for row in range(6, paper_plan.max_row + 1):
        if paper_plan.cell(row, 1).value == "Recommended manuscript":
            paper_plan.cell(row, 2).value = (
                "Leakage-aware and calibrated four-class brain MRI classification: "
                "internal validation, quantitative Grad-CAM analysis, and "
                "external-candidate reuse audits"
            )

    # Refresh the original diagnosis without erasing the historical gap text.
    audit_summary = wb["Audit Summary"]
    model_state = (
        "completed paired EfficientNet-B0/custom-CNN internal evaluation"
        if efficient and comparison
        else "completed custom-CNN internal evaluation with the EfficientNet-B0 run tracked separately"
    )
    audit_summary["D6"] = (
        "The repository now contains a leakage-controlled development manifest, "
        f"{model_state}, quantitative Grad-CAM analysis, candidate reuse audits, "
        "a data-quality report, and a preliminary paper-only RAG benchmark. The paper "
        "remains evidence-limited: neither audited external candidate was independent, "
        "patient identifiers are unavailable, and blinded XAI expert review is outstanding."
    )

    gap_status = {
        "Data leakage / source confounding": (
            "Complete",
            "Final 7,193-record manifest has no cross-split exact hash or duplicate/provenance component.",
        ),
        "Claim scope": (
            "Complete",
            "README, UI/report boundaries, protocol, model card and manuscript use research-only 2D dataset-label language.",
        ),
        "Missing empirical evidence": (
            "Partial" if not (efficient and comparison) else "Complete internally",
            (
                "Both internal models, paired comparison, XAI and RAG artifacts exist; external performance remains not estimable from rejected candidates."
                if efficient and comparison
                else "Custom CNN, XAI and RAG artifacts exist; EfficientNet/paired comparison is pending and external performance is not estimable from rejected candidates."
            ),
        ),
        "External validation": (
            "Blocked by evidence",
            "BRISC and BDNeuro-MRI v7 failed independence audits; a genuinely independent cohort is still required.",
        ),
        "Calibration and uncertainty": (
            "Complete for comparator",
            "Validation-only temperature scaling, ECE/Brier/NLL and risk–coverage are generated; custom-CNN scaling worsened test calibration.",
        ),
        "Statistical rigor": (
            "Complete internally" if comparison else "Complete for comparator",
            (
                "Grouped bootstrap confidence intervals, paired grouped-bootstrap differences and exact McNemar testing are generated for both models."
                if comparison
                else "Grouped bootstrap confidence intervals and the prespecified metric set are generated for the custom CNN."
            ),
        ),
        "XAI validation": (
            "Partial",
            "Repeatability, randomization sensitivity, IoU and pointing-game evaluation completed on 1,045 cases; blinded expert review remains.",
        ),
        "Tracker verification": (
            "Partial",
            "A normalized 74-record index exists; 9/10 shortlisted records received first-reviewer full-text extraction and one publisher/abstract verification. Second review remains.",
        ),
        "RAG evaluation": (
            "Preliminary",
            "A 10-case retrieval benchmark is complete; domain-expert review, citation precision and answer faithfulness remain.",
        ),
        "WHO CNS5 alignment": (
            "Complete",
            "All study-facing documents distinguish four source-dataset labels from integrated WHO CNS diagnoses.",
        ),
        "Fairness and subgroup analysis": (
            "Not estimable",
            "Age, sex, site and scanner metadata are absent; no subgroup or fairness claim is made.",
        ),
        "Reproducibility and reporting": (
            "Substantially complete",
            "Frozen protocol/config, manifests, hashes, data/model cards, environment snapshot, tests and machine-readable readiness checks exist.",
        ),
    }
    gaps = wb["Gap Matrix"]
    for merged_range in ["A1:G2", "A3:G3"]:
        if merged_range in {str(value) for value in gaps.merged_cells.ranges}:
            gaps.unmerge_cells(merged_range)
    gaps.merge_cells("A1:I2")
    gaps.merge_cells("A3:I3")
    gaps.cell(5, 8, "Execution status")
    gaps.cell(5, 9, "Evidence / remaining boundary")
    for col in range(1, 10):
        cell = gaps.cell(5, col)
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = border
    for row in range(6, gaps.max_row + 1):
        status, evidence = gap_status.get(
            str(gaps.cell(row, 2).value),
            ("Review", "See Research Status and machine-readable readiness audit."),
        )
        gaps.cell(row, 8, status)
        gaps.cell(row, 9, evidence)
        for col in range(1, 10):
            cell = gaps.cell(row, col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        gaps.cell(row, 8).fill = PatternFill(
            "solid",
            fgColor=(
                pale_green
                if status in {"Complete", "Complete internally", "Substantially complete", "Complete for comparator"}
                else pale_red
                if status in {"Blocked by evidence", "Not estimable"}
                else grey
            ),
        )
        gaps.row_dimensions[row].height = 78
    gaps.column_dimensions["H"].width = 24
    gaps.column_dimensions["I"].width = 58
    gaps.auto_filter.ref = f"A5:I{gaps.max_row}"

    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(WORKBOOK)
    print(WORKBOOK)


if __name__ == "__main__":
    main()
